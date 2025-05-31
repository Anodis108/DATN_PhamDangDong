from __future__ import annotations

import logging
import os
import time

import cv2
import numpy as np
import openpyxl
from PIL import ExifTags
from PIL import Image
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QPixmap
from PyQt5.QtWidgets import QDialog
from PyQt5.QtWidgets import QLabel
from PyQt5.QtWidgets import QVBoxLayout

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
)


class ImageUtils:
    @staticmethod
    def save_img(image: np.ndarray, save_dir: str) -> str:
        """Lưu ảnh numpy array thành file JPEG với timestamp."""
        timestamp = time.strftime('%Y%m%d_%H%M%S')
        filename = f'image_{timestamp}.jpg'
        save_path = os.path.join(save_dir, filename)
        cv2.imwrite(save_path, image)
        return save_path

    @staticmethod
    def show_image_window(img_path: str):
        """Hiển thị cửa sổ popup chứa ảnh từ đường dẫn."""
        window = QDialog()
        window.setWindowTitle('Hiển thị ảnh')
        window.resize(1440, 810)

        layout = QVBoxLayout()
        label = QLabel()
        pixmap = QPixmap(img_path)

        if pixmap.isNull():
            label.setText('Không thể đọc ảnh.')
        else:
            label.setPixmap(pixmap.scaled(1440, 810, Qt.KeepAspectRatio))

        label.setAlignment(Qt.AlignCenter)
        layout.addWidget(label)
        window.setLayout(layout)
        window.exec_()

    @staticmethod
    def save_to_excel(self, response, file_name='detection_data.xlsx'):
        # Kiểm tra xem file đã tồn tại chưa
        try:
            # Nếu file đã tồn tại, mở nó
            wb = openpyxl.load_workbook(file_name)
            sheet = wb.active
        except FileNotFoundError:
            # Nếu file chưa tồn tại, tạo file mới
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = 'Dữ liệu'
            # Tiêu đề cột nếu là file mới
            sheet.append([
                'Place', 'Height', 'Timestamp', 'Image Path',
            ])

        # Kiểm tra xem dữ liệu đã tồn tại dựa trên Place, Height, và Image Path
        data_exists = False
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=4):
            place_cell, height_cell, _, image_path_cell = row
            if (
                place_cell.value == response['place'] and
                str(height_cell.value) == str(response['height']) and
                image_path_cell.value == response['image_path']
            ):
                data_exists = True
                break

        if data_exists:
            logging.info(
                f'Dữ liệu Place: {response["place"]}, Height: {response["height"]}, Image Path: {response["image_path"]} đã tồn tại trong file Excel.',
            )
        else:
            # Thêm dữ liệu vào hàng mới nếu chưa có
            data = [
                response['place'],
                response['height'],
                response['timestamp'],
                response['image_path'],
            ]
            sheet.append(data)

            # Lưu vào file
            wb.save(file_name)
            logging.info(f'Dữ liệu đã được lưu vào {file_name}')

    @staticmethod
    def read_image_fix_orientation(path: str) -> np.ndarray:
        try:
            image = Image.open(path)

            # Sửa xoay nếu có EXIF
            try:
                for orientation in ExifTags.TAGS.keys():
                    if ExifTags.TAGS[orientation] == 'Orientation':
                        break
                exif = image._getexif()
                if exif is not None:
                    orientation_value = exif.get(orientation, None)
                    if orientation_value == 3:
                        image = image.rotate(180, expand=True)
                    elif orientation_value == 6:
                        image = image.rotate(270, expand=True)
                    elif orientation_value == 8:
                        image = image.rotate(90, expand=True)
            except Exception as e:
                print(f'[WARNING] Cannot correct orientation: {e}')

            return np.array(image.convert('RGB'))
        except Exception as e:
            print(f'[ERROR] Failed to read image {path}: {e}')
            return None
