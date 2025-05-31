from __future__ import annotations

import logging
from datetime import datetime

import cv2
import numpy as np
import openpyxl
from common.utils import get_settings
from main_controller.utils import ImageUtils
from model import BaseResults
from network.call_api import APICaller
from PIL import ExifTags
from PIL import Image
from PyQt5.QtCore import pyqtSignal
from PyQt5.QtCore import QThread
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
)
settings = get_settings()


class ThreadWork(QThread):
    image_processed = pyqtSignal(object, str, float, str)
    video_frame_processed = pyqtSignal(np.ndarray)
    error_occurred = pyqtSignal(str)
    finished = pyqtSignal()

    def __init__(self, image_path, is_video):
        super().__init__()
        self.image_path = image_path
        self.is_video = is_video
        self.running = True
        self.video_capture = None
        self.callapi = APICaller()
        self.img_utils = ImageUtils()

    def run(self):
        if self.is_video or self.image_path == '0':
            self.run_video_stream()
        else:
            self.run_image_processing()
        self.finished.emit()

    def run_image_processing(self):
        # 📌 1. Đọc ảnh gốc với Pillow và chỉnh orientation
        try:
            image_pil = Image.open(self.image_path)
            # Auto-rotate theo EXIF
            try:
                for orientation in ExifTags.TAGS.keys():
                    if ExifTags.TAGS[orientation] == 'Orientation':
                        break
                exif = image_pil._getexif()
                if exif is not None:
                    orientation_value = exif.get(orientation, None)
                    if orientation_value == 3:
                        image_pil = image_pil.rotate(180, expand=True)
                    elif orientation_value == 6:
                        image_pil = image_pil.rotate(270, expand=True)
                    elif orientation_value == 8:
                        image_pil = image_pil.rotate(90, expand=True)
            except Exception as e:
                print(
                    f'[Warning] No EXIF orientation or failed to rotate: {e}',
                )
        except Exception as e:
            self.error_occurred.emit(f'Failed to load image: {e}')
            return

        # 📌 2. Chuyển đổi sang NumPy để xử lý (cv2 hoặc model inference)
        image_np = np.array(image_pil.convert('RGB'))  # RGB → ndarray
        image_bgr = cv2.cvtColor(image_np, cv2.COLOR_RGB2BGR)

        # 📌 3. Lưu ảnh ban đầu
        save_img_path = self.img_utils.save_img(
            image_bgr, settings.app.save_dir,
        )

        # 📌 4. Gọi API tính chiều cao
        self.response: BaseResults = self.callapi.call_api(
            settings.host_height_service, image_bgr, self.image_path,
        )
        self.height = max(self.response.heights)

        # 📌 5. Đọc lại ảnh đã xử lý đầu ra bằng Pillow
        image_pil2 = Image.open(self.response.out_path).convert('RGB')

        # 📌 6. Lưu vào Excel
        self.save_to_excel(
            {
                'place': 'HaUI',
                'height': self.height,
                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'image_path': save_img_path,
            }, 'detection_data.xlsx',
        )

        # 📌 7. Emit ảnh (RGB numpy array) cho PyQt
        self.image_processed.emit(
            image_pil2, 'HaUI', self.height, self.response.out_path,
        )

    def run_video_stream(self):
        self.video_capture = cv2.VideoCapture(self.image_path)

    def process_video_frame(self):
        if self.video_capture and self.video_capture.isOpened() and self.running:
            ret, frame = self.video_capture.read()
            if ret:
                rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                self.video_frame_processed.emit(rgb_frame)
            else:
                self.running = False
                self.finished.emit()

    def save_to_excel(self, response, file_name='detection_data.xlsx'):
        try:
            # Kiểm tra xem file đã tồn tại chưa
            try:
                wb = openpyxl.load_workbook(file_name)
                sheet = wb.active
            except FileNotFoundError:
                wb = openpyxl.Workbook()
                sheet = wb.active
                sheet.title = 'Dữ liệu'
                sheet.append(['Place', 'Height', 'Timestamp', 'Image Path'])

            # Kiểm tra xem dữ liệu đã tồn tại dựa trên Place, Height, và Image Path
            data_exists = False
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=4):
                place_cell, height_cell, _, image_path_cell = row
                if (
                    place_cell.value == response['place'] and
                    str(height_cell.value) == str(response['height']) and
                    image_path_cell.value == response['image_path']
                ):
                    data_exists = True
                    break

            if data_exists:
                logging.info(
                    f'Dữ liệu Place: {response["place"]}, Height: {response["height"]}, Image Path: {response["image_path"]} đã tồn tại trong file Excel.',
                )
            else:
                data = [
                    response['place'],
                    response['height'],
                    response['timestamp'],
                    response['image_path'],
                ]
                sheet.append(data)
                wb.save(file_name)
                logging.info(f'Dữ liệu đã được lưu vào {file_name}')
        except Exception as e:
            logging.error(f'Error saving to Excel: {e}')
            self.error_occurred.emit(f'Error saving to Excel: {e}')

    def stop(self):
        self.running = False
        if self.video_capture:
            self.video_capture.release()
        self.quit()
        self.wait()
